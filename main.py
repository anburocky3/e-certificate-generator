from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._-") or "student"


def normalize_name(value: str) -> str:
    value = re.sub(r"\s+", " ", value.strip())
    if not value:
        return value

    def capitalize_chunk(chunk: str) -> str:
        if not chunk:
            return chunk
        return chunk[0].upper() + chunk[1:].lower()

    def normalize_token(token: str) -> str:
        trailing_dot = token.endswith(".")
        core = token[:-1] if trailing_dot else token

        # Handle dotted initials like r.k. / r.k / a.b.c
        if "." in core:
            parts = [p for p in core.split(".") if p]
            if parts and all(len(p) == 1 and p.isalpha() for p in parts):
                result = ".".join(p.upper() for p in parts)
                if trailing_dot:
                    result += "."
                return result

        # Handle single-letter initial like r / r.
        if len(core) == 1 and core.isalpha():
            result = core.upper()
            if trailing_dot:
                result += "."
            return result

        # Capitalize regular names while preserving hyphen/apostrophe separators.
        pieces = re.split(r"([-'])", core)
        normalized_pieces = [capitalize_chunk(piece) if piece not in {"-", "'"} else piece for piece in pieces]
        result = "".join(normalized_pieces)
        if trailing_dot:
            result += "."
        return result

    return " ".join(normalize_token(token) for token in value.split(" "))


def parse_color(value: str) -> tuple[int, int, int]:
    value = value.strip().lstrip("#")
    if len(value) != 6:
        raise ValueError("Color must be a 6-digit hex value, e.g. 1A1A1A")
    red = int(value[0:2], 16)
    green = int(value[2:4], 16)
    blue = int(value[4:6], 16)
    return red, green, blue


def load_students(
    excel_path: Path,
    sheet_name: str | None,
    name_column: str,
    start_row: int | None,
    end_row: int | None,
) -> list[dict[str, str]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel file has no header row.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    if name_column not in headers:
        raise ValueError(f"Column '{name_column}' not found. Available columns: {headers}")

    name_index = headers.index(name_column)
    students: list[dict[str, str]] = []

    if start_row is not None and start_row < 2:
        raise ValueError("--start-row must be 2 or greater (row 1 is the header).")
    if end_row is not None and end_row < 2:
        raise ValueError("--end-row must be 2 or greater (row 1 is the header).")
    if start_row is not None and end_row is not None and start_row > end_row:
        raise ValueError("--start-row must be less than or equal to --end-row.")

    iter_start = start_row if start_row is not None else 2
    iter_end = end_row

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=iter_start, max_row=iter_end, values_only=True), start=iter_start
    ):
        if not row:
            continue
        raw_name = row[name_index] if name_index < len(row) else None
        if raw_name is None:
            continue
        name = normalize_name(str(raw_name))
        if not name:
            continue
        students.append({"name": name, "row": str(row_number)})

    return students


def resolve_font(font_path: Path | None, font_size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    if font_path:
        return ImageFont.truetype(str(font_path), font_size)

    # Prefer Pacifico when no explicit font path is passed.
    pacifico_candidates = [
        Path.cwd() / "assets" / "fonts" / "Pacifico-Regular.ttf",
        Path.cwd() / "Pacifico-Regular.ttf",
        Path("C:/Windows/Fonts/Pacifico.ttf"),
        Path("C:/Windows/Fonts/Pacifico-Regular.ttf"),
    ]
    for candidate in pacifico_candidates:
        if candidate.exists():
            try:
                return ImageFont.truetype(str(candidate), font_size)
            except OSError:
                pass

    try:
        return ImageFont.truetype("arial.ttf", font_size)
    except OSError:
        return ImageFont.load_default()


def draw_centered_name(
    image: Image.Image,
    name: str,
    font: ImageFont.ImageFont,
    color: tuple[int, int, int],
    x: int | None,
    y: int | None,
    zone_top_ratio: float,
    zone_bottom_ratio: float,
    zone_padding: int,
) -> None:
    draw = ImageDraw.Draw(image)
    left, top, right, bottom = draw.textbbox((0, 0), name, font=font)
    text_width = int(right - left)
    text_height = int(bottom - top)

    text_x = x if x is not None else (image.width - text_width) // 2

    if y is not None:
        text_y = y
    else:
        # Place the name inside a configurable vertical zone (between heading and paragraph).
        zone_top = int(image.height * zone_top_ratio) + zone_padding
        zone_bottom = int(image.height * zone_bottom_ratio) - zone_padding

        if zone_bottom <= zone_top:
            text_y = (image.height - text_height) // 2
        else:
            target_y = zone_top + ((zone_bottom - zone_top - text_height) // 2)
            min_y = int(max(zone_top, 0))
            max_y = int(max(min_y, zone_bottom - text_height))
            text_y = int(min(max(target_y, min_y), max_y))

    draw.text((text_x, text_y), name, fill=color, font=font)


def generate_certificates(args: argparse.Namespace) -> None:
    excel_path = Path(args.excel).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    font_path = Path(args.font).expanduser().resolve() if args.font else None

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template image not found: {template_path}")
    if font_path and not font_path.exists():
        raise FileNotFoundError(f"Font file not found: {font_path}")
    if not 0 <= args.name_zone_top <= 1 or not 0 <= args.name_zone_bottom <= 1:
        raise ValueError("--name-zone-top and --name-zone-bottom must be between 0 and 1.")
    if args.name_zone_top >= args.name_zone_bottom:
        raise ValueError("--name-zone-top must be less than --name-zone-bottom.")
    if args.name_zone_padding < 0:
        raise ValueError("--name-zone-padding must be non-negative.")

    color = parse_color(args.color)
    students = load_students(excel_path, args.sheet, args.name_column, args.start_row, args.end_row)
    if args.limit:
        students = students[: args.limit]

    if not students:
        print("No students found in the Excel file.")
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    font = resolve_font(font_path, args.font_size)

    generated = 0
    skipped = 0

    for index, student in enumerate(students, start=1):
        name = student["name"]
        safe_name = sanitize_filename(name)
        output_file = output_dir / f"{args.prefix}_{index:03d}_{safe_name}.png"

        if output_file.exists() and not args.overwrite:
            skipped += 1
            continue

        if args.dry_run:
            print(f"[DRY RUN] {name} -> {output_file.name}")
            generated += 1
            continue

        with Image.open(template_path).convert("RGB") as image:
            draw_centered_name(
                image,
                name,
                font,
                color,
                args.name_x,
                args.name_y,
                args.name_zone_top,
                args.name_zone_bottom,
                args.name_zone_padding,
            )
            image.save(output_file, format="PNG")

        generated += 1

    range_start = args.start_row if args.start_row is not None else 2
    range_end = args.end_row if args.end_row is not None else "last"
    print(
        f"Done. Generated: {generated}, Skipped: {skipped}, Total rows read: {len(students)}, "
        f"Excel rows processed: {range_start} to {range_end}"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate e-certificates from Excel data.")
    parser.add_argument("--excel", default="input/data.xlsx", help="Path to Excel file (.xlsx).")
    parser.add_argument("--template", default="input/template.png", help="Path to certificate template image.")
    parser.add_argument("--output-dir", default="output", help="Directory for generated certificates.")
    parser.add_argument("--sheet", default=None, help="Optional sheet name. Defaults to active sheet.")
    parser.add_argument("--name-column", default="Name", help="Column header containing student names.")
    parser.add_argument("--start-row", type=int, default=None, help="First Excel row to process (inclusive, row 1 is header).")
    parser.add_argument("--end-row", type=int, default=None, help="Last Excel row to process (inclusive).")
    parser.add_argument("--font", default=None, help="Optional .ttf/.otf font file path.")
    parser.add_argument("--font-size", type=int, default=40, help="Font size for the student name.")
    parser.add_argument("--name-x", type=int, default=None, help="Fixed X position. Default is centered.")
    parser.add_argument("--name-y", type=int, default=None, help="Fixed Y position. Overrides zone placement when set.")
    parser.add_argument(
        "--name-zone-top",
        type=float,
        default=0.44,
        help="Top of vertical placement zone as image-height ratio (default: 0.44).",
    )
    parser.add_argument(
        "--name-zone-bottom",
        type=float,
        default=0.58,
        help="Bottom of vertical placement zone as image-height ratio (default: 0.58).",
    )
    parser.add_argument(
        "--name-zone-padding",
        type=int,
        default=20,
        help="Extra inner padding in pixels for the vertical placement zone.",
    )
    parser.add_argument("--color", default="FFFFFF", help="Name text color as hex, e.g. FFFFFF.")
    parser.add_argument("--prefix", default="certificate", help="Output file prefix.")
    parser.add_argument("--limit", type=int, default=None, help="Optional row limit for quick testing.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files.")
    parser.add_argument("--dry-run", action="store_true", help="Preview file names without writing images.")

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    generate_certificates(args)


if __name__ == "__main__":
    main()
