from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .utils import normalize_key, normalize_name, normalize_lookup_value


@dataclass(slots=True)
class StudentRecord:
    name: str
    row: int
    roll_no: str | None = None
    email: str | None = None


def _find_header_index(headers: list[str], target: str | None, aliases: tuple[str, ...]) -> int | None:
    if target:
        normalized_target = normalize_key(target)
        for index, header in enumerate(headers):
            if normalize_key(header) == normalized_target:
                return index

    for alias in aliases:
        normalized_alias = normalize_key(alias)
        for index, header in enumerate(headers):
            if normalize_key(header) == normalized_alias:
                return index

    return None


def _normalize_optional_cell(value: object, *, lower: bool = False) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.lower() if lower else text


def load_students(
    excel_path: Path,
    sheet_name: str | None,
    name_column: str,
    roll_column: str | None,
    email_column: str | None,
    start_row: int | None,
    end_row: int | None,
) -> list[StudentRecord]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel file has no header row.")

    headers = [str(header).strip() if header is not None else "" for header in header_row]
    name_index = _find_header_index(headers, name_column, ("Name",))
    if name_index is None:
        raise ValueError(f"Column '{name_column}' not found. Available columns: {headers}")

    roll_index = _find_header_index(
        headers,
        roll_column,
        (
            "Roll No",
            "RollNo",
            "Roll Number",
            "Roll_Number",
            "roll_no",
        ),
    )
    email_index = _find_header_index(headers, email_column, ("Email", "E-mail", "Email Address", "mail"))

    if start_row is not None and start_row < 2:
        raise ValueError("--start-row must be 2 or greater (row 1 is the header).")
    if end_row is not None and end_row < 2:
        raise ValueError("--end-row must be 2 or greater (row 1 is the header).")
    if start_row is not None and end_row is not None and start_row > end_row:
        raise ValueError("--start-row must be less than or equal to --end-row.")

    iter_start = start_row if start_row is not None else 2
    iter_end = end_row
    students: list[StudentRecord] = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=iter_start, max_row=iter_end, values_only=True),
        start=iter_start,
    ):
        if not row:
            continue

        raw_name = row[name_index] if name_index < len(row) else None
        if raw_name is None:
            continue

        name = normalize_name(str(raw_name))
        if not name:
            continue

        roll_no = _normalize_optional_cell(row[roll_index]) if roll_index is not None and roll_index < len(row) else None
        email = (
            normalize_lookup_value(row[email_index])
            if email_index is not None and email_index < len(row)
            else None
        )

        students.append(
            StudentRecord(
                name=name,
                row=row_number,
                roll_no=roll_no,
                email=email,
            )
        )

    return students

